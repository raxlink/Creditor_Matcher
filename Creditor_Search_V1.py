import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from typing import List, Tuple
import os
class CreditorSearchGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Creditor Search Tool")
        self.SHEET_NAME = "Creditors"
        self.NEW_SHEET_NAME = "Matched Results"  # New sheet to store matched results
        
        # Set minimum window size
        self.root.minsize(400, 200)
        
        # Center window on screen
        window_width = 400
        window_height = 200
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        self.root.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')
        
        self.create_widgets()

    def create_widgets(self):
        # Create main frame
        main_frame = tk.Frame(self.root, padx=20, pady=20)
        main_frame.pack(expand=True, fill='both')

        # Add title label
        title_label = tk.Label(
            main_frame, 
            text="Creditor Search Tool",
            font=('Helvetica', 14, 'bold')
        )
        title_label.pack(pady=(0, 20))

        # Add description
        description = tk.Label(
            main_frame,
            text=f"This tool will search Column E in the '{self.SHEET_NAME}' sheet\nand highlight matches in yellow.",
            justify=tk.CENTER
        )
        description.pack(pady=(0, 20))

        # Add select file button
        select_button = tk.Button(
            main_frame,
            text="Select Excel File",
            command=self.process_file,
            width=20,
            height=2
        )
        select_button.pack()

    def process_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls")],
            initialdir=os.path.expanduser("~")
        )
        
        if file_path:
            try:
                matches, matched_names = self.search_creditors(file_path)
                self.show_results(matches, matched_names)
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def get_company_names(self) -> list[str]:
        # (Same list of company names as before)
        return [# subNames1
        "Alice and Olivia", "Alice & Olivia", "Allegion Americas", "Schlage Lock", "AlphaTheta", "Alpha Theta", 
        "Bar-S Foods", "Campofrio", "Sigma International", "Bella + Canvas", "Alo", "Ingrid & Isabel", 
        "Bella Corp", "Bella Corporation", "Blue Triton", "Nestle Water", "Brown & Haley", "Brown and Haley", 
        "Caleres", "Brown Shoe", "Carolina Turkeys", "Butterball", "Cavendish Farms", "Irving Tissue", 
        "Midland Transport", "Chattem", "Sanofi", "Citrus World", "Florida's Natural", "Danone Business Services Finance", 
        "Dannon", "Delta Faucet", "Peerless Faucet", "Masco", "Ecolab", "Nalco", "EDC", "Export Development Canada", 
        "EightMood", "Eight Mood", "Essex Manufacturing", "Baum Essex", "Fuji Photo", "Fuji Film", "Fujifilm", 
        "General Mills", "Pillsbury", "H.O. Sports", "HO Sports", "H O Sports", "Heathcote & Ivory", "Heathcote", 
        "HH Brown Shoe", "Justin Brands", "High Liner Foods", "Icelandic", "Huhtamaki", "Chinet", "Idahoan Foods", "IFP", 
        "InRoad Toys", "In Road Toys", "Johnson & Johnson", "J&J", "Keeco", "Hollander Sleep", "Kent Water Sports", 
        "Kent Sporting Goods", "KidSource", "Kid Source", "Lew's Fishing", "Lews Fishing", "Marc Fisher", "MODA", 
        "Mattel", "Fisher Price", "McCormick and Company", "McCormick and Co", "McCormick & Company", "McCormick & Co", 
        "MGA Entertainment", "Micro Games", "MI Windows & Doors", "MI Windows and Doors", "National Beverage", 
        "NewBevCo", "Novolex", "Bagcraft", "Oil-Dri", "Oil Dry", "Panasonic", "Matsushita", "Paradign8697", "Paradigm 8697", 
        "Pepsi", "Frito Lay", "Performance Food", "Performance Foodservice", "Phillips Van Heusen", "PVH", 
        "Procter & Gamble", "Procter and Gamble", "P&G", "Purdue Pharma", "Rhodes Pharmaceutical", "Purity Wholesale Grocers", 
        "Supreme Distributors", "R.G. Barry", "RG Barry", "Reckitt", "Mead Johnson", "Retail Consulting Services", 
        "RCS Real Estate Advisors", "RSI", "Restaurant Services", "RXBar", "RX Bar", "Sigma International", "Sigma Intl", 
        "Smithfield", "Farmland Foods", "SnapToys", "Snap Toys", "Star Children's Dress", "Rare Editions For Girls", 
        "Strike King", "Striking Lures", "Sturm", "Ruger", "T. Marzetti", "T Marzetti", "T.Marzetti", "Topgolf", "Callaway", 
        "Unibail-Rodamco-Westfield", "Westfield", "VF Corp", "Jansport", "Voxx", "Audiovox", "W.C. Bradley", "WC Bradley", 
        "W.L. Gore", "WL Gore", "Water Innovations", "Moen", "WeCool Toys", "We Cool Toys", "Wells Enterprises", "Blue Bunny", 
        "William Carter Company", "William Carter Co", "Williamson-Dickie", "Williamson Dickie", "WM Wrigley", "W.M. Wrigley", 
        "Wonderful Pistachios & Almonds", "Wonderful Pistachios",

        # subNames2
        "7 For All Mankind", "A360 Media", "ACH Food", "Acme United", "Acushnet", "Ajinomoto Foods", "Alexander Dolls", 
        "Alliance Consumer Group", "Amer Sports", "American Bank Note", "American Greetings", "AMYNTA", "Anacapri Foods", 
        "Anda", "Anova Foods", "Antigua Group", "Arthur Schuman", "Atlas Roofing", "Attwood", "Audio Technica", "B Riley Financial", 
        "B&G Foods", "BABY BRIEFCASE", "Baggallini", "BakeMark", "Balvi Gifts", "Banana Panda", "Bandai", "Barilla", 
        "Barnwood Living", "Barrett Firearms", "Basic Fun", "Bay Valley Foods", "Bayer Consumer Care", "Bayer HealthCare", 
        "Bayer Pharmaceuticals", "BE AMAZING", "BearPaw", "Beatriz Ball", "Beiersdorf", "Ben E. Keith", "Bethany Lowe Designs", 
        "Bic", "Big Geyser", "Big Rock Sports", "BIGJIGS TOYS", "Bimbo Bakeries", "Birkenstock", "Bissell", "Bolthouse Farms", 
        "BooKid, TomRide", "BOPS", "Bozzuto's", "Brooks Sports", "BSH Home Appliances", "Bubblegum Stuff", "Buddy & Barney", 
        "Bumble Bee Seafood", "Burton Snowboards", "Bush Brothers", "Busseto Foods", "C&S Wholesale Grocers", "Cabot Creamery", 
        "Calypso", "Campbell Soup", "Candide Baby", "Cargill", "Carhartt", "Carlisle Construction", "Cascade Designs", "Casio", 
        "Catelli Foods", "Chamberlain Group", "Chaney Instruments", "Charbroil", "Church & Dwight", "Citigroup Global Market", 
        "Citizen Watch", "Citizen Watch Group", "Citterio", "CITYPETS", "CL Products", "Clarks", "Clorox Sales", "CO Lynch", 
        "Cobra Puma Golf", "Coca-Cola", "Coca-Cola", "Coca-Cola", "Coca-Cola", "Columbia Sportswear", "Commercial Credit Consultants", 
        "ConAgra", "Cookies United", "Core-Mark", "Corolle", "Coty", "Crescent Packing", "Crow Holdings Capita", "D&H Distributing", 
        "Dairy Farmers", "Daisy", "DAM", "Darigold", "David Oppenheimer", "Dawn Food", "Dean Foods", "Dearforms", 
        "Deckers Outdoor", "Del Monte", "Del Monte", "Delsey Luggage", "Delta Apparel", "Deutsche Bank", "DF Stauffer Biscuit", 
        "di Caro", "DJECO", "DOMESTIC MARKETING, Yookidoo", "Domino Sugar", "Dot Foods", "Duracell", "Dyno Merchandise", 
        "Eagle Family Foods", "Ecco", "Edgewell", "Efco Products", "Elleven", "Elmer's Chocolate", "Energizer", "Escalade Sports", 
        "Estee Lauder", "Exxel Outdoors", "Faribault Foods", "Ferrara Candy", "Ferrero", "Fieldale Farms", "Fikkerts", "Fila", 
        "Find Your Glow", "Fizz Creations", "Formation Brands", "Frankford Candy", "Franklin Sports", "Fratelli Beretta", 
        "Freaker USA", "Free Motion Fitness", "Fresh Mark", "Fruit of The Loom", "Funko", "Galderma Laboratories", "GALT TOYS", 
        "Garan Manufacturing", "Genesco", "GEOMAGWORLD", "George Delallo", "Getzler Henrich", "Gibson Games", "G-III Apparel", 
        "GlaxoSmithKline", "GOLDEN RABBIT II", "Gordon Food", "Gourmet Boutique", "Goya Foods", "Great Lakes Cheese", 
        "Greater Omaha Packing", "Gregory Mountain", "Groupe SEB", "GSI Outdoors", "GUESS", "Haleon", "Hamilton Beach", 
        "Hampshire Companies", "Handbag Butler", "Hartz Mountain", "Haywire Group", "Head/Penn Racquet", "Helen of Troy", 
        "Hormel Foods", "House of RaeFord Farms", "Hull Property", "Hunter Walton", "Iconex", "IDB Factors", "Ingram Content Group", 
        "Inline Plastics", "Intex Recreation", "Invesco", "JAECI", "Jasco Products", "Johnson Outdoors", "Johnvince Foods", 
        "JSO Associates", "Junk Gypsy", "Just Born", "Keen Footwear", "Keen-Summit", "Kellanova", "Kellogg", "Kellwood", 
        "Ken's Foods", "Kenvue", "Kids Toy Club", "Kidz Delight", "Kidzaw", "Kikkoman", "Kimco Realty", "Knouse Foods", 
        "Kraft Heinz", "Kreider Farms", "Kruger Products", "K-Swiss", "LaCrosse Footwear", "Lamb Weston", "Lasso", "Lavazza", 
        "Lay's", "LC Industries", "Lego", "Linden", "Lionshead Apparel", "Lion's Head", "Little Brownie Bakers", "Little Tikes", 
        "LiveNation", "Lulus", "Luna", "Lyle Industries", "Mad Engine", "Maguire Group", "Manitowoc", "Mann Packing", 
        "Maruchan", "Mars", "Mattel", "Mayco", "McCormick", "Merriam-Webster", "Metra Electronics", "Michael Kors", "Miele", 
        "Mills Fleet Farm", "Mizuno USA", "MLB", "Munchkin", "Nabisco", "Namco Bandai", "Nautica", "Nestle Waters North America", 
        "Nike", "Nikon", "Nissin Foods", "Noble House", "Nordstrom", "Oasis Brands", "Olin Corporation", "Olivier Group", 
        "Ollie's Bargain Outlet", "Omega Engineering", "Oneida", "OppenheimerFunds", "Orion", "Pactiv", "Pall Corporation", 
        "Panasonic", "Parker Hannifin", "Pella", "PepsiCo", "PetSmart", "Philips Electronics", "Playtex", "Procter & Gamble", 
        "Purdue Foods", "RCA", "Reckitt", "Rembrandt", "Rhino-Rack", "Ricola", "Riedel", "Rosenthal", "S.H. Heilig", 
        "S.C. Johnson", "SC Johnson", "Sears", "SharkNinja", "Simplay3", "Skechers", "Smith & Wesson", "Smithfield Foods", 
        "Spectrum Brands", "Stanley Black & Decker", "Sunbeam", "Sunglass Hut", "Swisher", "T-Mobile", "TAPCO", "The Honest Company",
        "The Macallan", "The North Face", "Toys R Us", "Tupperware", "Tyson Foods", "United Natural Foods", "Vanguard Group", 
        "Vans", "Verizon", "Volkswagen", "Walmart", "Wells Fargo", "Wesfarmers", "Whirlpool", "Xerox", "Zara", "Zoku"
        ]

    def search_creditors(self, filepath: str) -> tuple[int, list[str]]:
        """Search for company names in the specified Excel sheet and highlight matches."""
        try:
            wb = load_workbook(filepath)
        except Exception as e:
            raise Exception(f"Error opening Excel file:\n{str(e)}")
        
        # Check for sheet existence with better error message
        if self.SHEET_NAME not in wb.sheetnames:
            available_sheets = ", ".join(wb.sheetnames)
            raise Exception(
                f"Sheet '{self.SHEET_NAME}' not found in workbook.\n\n"
                f"Available sheets are: {available_sheets}"
            )
        
        ws = wb[self.SHEET_NAME]
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        company_names = self.get_company_names()
        match_count = 0
        matched_companies = []
        matched_rows = []

        # Search column E and collect matches
        try:
            for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):  # Start from row 2 (skip header)
                for cell in row:
                    if cell.value:  # Skip empty cells
                        cell_value = str(cell.value).strip()
                        if cell_value in company_names:
                            match_count += 1
                            matched_companies.append(cell_value)
                            matched_rows.append(cell.row)  # Collect the matched row number
                            cell.fill = yellow_fill
        except Exception as e:
            raise Exception(f"Error processing Column E:\n{str(e)}")

        # Create a new sheet for matched results
        if self.NEW_SHEET_NAME in wb.sheetnames:
            new_ws = wb[self.NEW_SHEET_NAME]  # If sheet exists, clear it
            new_ws.delete_rows(2, new_ws.max_row)  # Clear existing data in the sheet (excluding header)
        else:
            new_ws = wb.create_sheet(self.NEW_SHEET_NAME)  # Create a new sheet

        # Copy matched rows to the new sheet
        if matched_rows:
            header = ws[1]  # Assuming header is in the first row
            for col_idx, header_cell in enumerate(header, start=1):
                new_ws.cell(row=1, column=col_idx).value = header_cell.value  # Copy header to new sheet
            
            for i, row_num in enumerate(matched_rows, start=2):
                for col_idx, cell in enumerate(ws[row_num], start=1):
                    new_ws.cell(row=i, column=col_idx).value = cell.value  # Copy the entire matched row

        # Save the workbook
        try:
            wb.save(filepath)
        except PermissionError:
            raise Exception(
                "Unable to save the Excel file. Please ensure that:\n"
                "1. The file is not open in Excel\n"
                "2. You have write permissions for this file\n"
                "3. The file is not read-only"
            )
        except Exception as e:
            raise Exception(f"Error saving Excel file:\n{str(e)}")
        
        return match_count, matched_companies

    def show_results(self, matches: int, matched_names: list[str]):
        """Show results in a new window"""
        results_window = tk.Toplevel(self.root)
        results_window.title("Search Results")
        
        # Center the results window
        window_width = 400
        window_height = 300
        screen_width = results_window.winfo_screenwidth()
        screen_height = results_window.winfo_screenheight()
        center_x = int(screen_width/2 - window_width/2)
        center_y = int(screen_height/2 - window_height/2)
        results_window.geometry(f'{window_width}x{window_height}+{center_x}+{center_y}')

        # Create main frame with padding
        frame = tk.Frame(results_window, padx=20, pady=20)
        frame.pack(expand=True, fill='both')

        # Add results header
        header = tk.Label(
            frame,
            text=f"Found {matches} matches in Column E",
            font=('Helvetica', 12, 'bold')
        )
        header.pack(pady=(0, 10))

        if matches > 0:
            # Create text widget for matches
            text_widget = tk.Text(frame, height=10, width=40)
            text_widget.pack(pady=(0, 10))
            
            # Add scrollbar
            scrollbar = tk.Scrollbar(frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            
            # Configure scrollbar
            text_widget.config(yscrollcommand=scrollbar.set)
            scrollbar.config(command=text_widget.yview)
            
            # Insert matched names
            for i, name in enumerate(matched_names, 1):
                text_widget.insert(tk.END, f"{i}. {name}\n")
            
            text_widget.config(state=tk.DISABLED)  # Make text read-only

        # Add status message
        status = tk.Label(
            frame,
            text="Matches have been highlighted in yellow in the Excel file.\n"
                 f"And copied to the '{self.NEW_SHEET_NAME}' sheet.",
            wraplength=350
        )
        status.pack(pady=(10, 0))

    def run(self):
        """Start the GUI application"""
        self.root.mainloop()

if __name__ == "__main__":
    app = CreditorSearchGUI()
    app.run()
